import time
from uuid import uuid4
from flask.helpers import send_from_directory
from openpyxl.styles.borders import Side
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Alignment
import openpyxl
import requests
from settings import File_UPLOAD_PATH
import os
from flask import Blueprint
from flask.templating import render_template
from ..exts.forms import ExcelForm

qrcode_bp=Blueprint("qrcode",__name__)

@qrcode_bp.route("/generate_qrcode",methods=["GET","POST"])
def generate_qrcode():
    excel_form=ExcelForm()
    if excel_form.validate_on_submit():
        f=excel_form.file.data
        filename=uuid4().hex+os.path.splitext(f.filename)[1]
        upload_path = os.path.join(
            File_UPLOAD_PATH, filename)
        f.save(upload_path)
        # 用openpyxl打开文件 检查文件是否符合要求
        wb_read=openpyxl.load_workbook(upload_path)
        sheet = wb_read.worksheets[0]
        if sheet.max_column==5:
            row_list = []
            for row in sheet.rows:
                col_list = []
                for i in range(len(row)):
                    col_list.append(row[i].value)
                row_list.append(col_list)
            print(row_list)
            for i in range(len(row_list)-1):
                text = row_list[0][0]+":"+row_list[i+1][0]+","+row_list[0][1]+":"+row_list[i+1][1]+","+row_list[0][2]+":" + \
                    str(row_list[i+1][2])+","+row_list[0][3]+":" + \
                    str(row_list[i+1][3])+","+row_list[0][4]+":"+row_list[i+1][4]
                # print(text)
                # 根据文字内容生成二维码图片
                params = {"data": text}
                time.sleep(0.5)
                content = requests.get(
                    url="https://api.qrserver.com/v1/create-qr-code/?size=150x150&", params=params).content
                with open(File_UPLOAD_PATH+os.sep+"二维码"+str(i+1)+".png", "wb") as f:
                    f.write(content)
            # 图片插入excel，文字写入一个新的excel
            wb = openpyxl.Workbook()
            # wb.create_sheet("带二维码")
            sheet = wb.active
            # 设置列宽
            sheet.column_dimensions["A"].width = 19.63
            sheet.column_dimensions["B"].width = 18.25
            sheet.column_dimensions["C"].width = 45
            # 设置边框
            border = Border(left=Side(style='thin', color='FF110000'), right=Side(style='thin', color='FF110000'), top=Side(style='thin', color='FF110000'), bottom=Side(style='thin', color='FF110000'), diagonal=Side(
                style='thin', color='FF000000'), diagonal_direction=0, outline=Side(style='thin', color='FF000000'), vertical=Side(style='thin', color='FF000000'), horizontal=Side(style='thin', color='FF000000'))
            j = 0
            for i in range(len(row_list)-1):
                # 设置行高第1 7行
                sheet.row_dimensions[j+1].height = 22.5
                sheet.cell(row=j+1, column=2).value = row_list[0][0]
                sheet.cell(
                    row=j+1, column=2).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+1, column=2).border = border
                # 二维码所在单元格的边框
                sheet.cell(row=j+1, column=1).border = border
                sheet.cell(
                    row=j+1, column=1).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+1, column=3).value = row_list[i+1][0]
                sheet.cell(
                    row=j+1, column=3).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+1, column=3).border = border
                # 设置行高第2 8行
                sheet.row_dimensions[j+2].height = 22.5
                sheet.cell(row=j+2, column=2).value = row_list[0][1]
                sheet.cell(
                    row=j+2, column=2).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+2, column=2).border = border
                sheet.cell(row=j+2, column=3).value = row_list[i+1][1]
                sheet.cell(
                    row=j+2, column=3).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+2, column=3).border = border
                # 设置行高第3 9行
                sheet.row_dimensions[j+3].height = 22.5
                sheet.cell(row=j+3, column=2).value = row_list[0][2]
                sheet.cell(
                    row=j+3, column=2).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+3, column=2).border = border
                sheet.cell(row=j+3, column=3).value = row_list[i+1][2]
                sheet.cell(
                    row=j+3, column=3).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+3, column=3).border = border

                sheet.row_dimensions[j+4].height = 22.5
                sheet.cell(row=j+4, column=2).value = row_list[0][3]
                sheet.cell(
                    row=j+4, column=2).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+4, column=2).border = border
                sheet.cell(row=j+4, column=3).value = row_list[i+1][3]
                sheet.cell(
                    row=j+4, column=3).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+4, column=3).border = border

                sheet.row_dimensions[j+5].height = 22.5
                sheet.cell(row=j+5, column=2).value = row_list[0][4]
                sheet.cell(
                    row=j+5, column=2).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+5, column=2).border = border
                sheet.cell(row=j+5, column=3).value = row_list[i+1][4]
                sheet.cell(
                    row=j+5, column=3).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+5, column=3).border = border

                sheet.row_dimensions[j+6].height = 22.5
                sheet.cell(row=j+6, column=2).value = ""
                sheet.cell(
                    row=j+6, column=2).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=j+6, column=3).value = ""
                sheet.cell(
                    row=j+6, column=3).alignment = Alignment(horizontal='center', vertical='center')

                # 合并单元格
                sheet.merge_cells(start_row=j+1, start_column=1, end_row=j+5, end_column=1)
                # 插入二维码
                sheet.add_image(Image(File_UPLOAD_PATH+os.sep+"二维码"+str(i+1)+".png"), "A"+str(j+1))
                j = j+6
            newfilename=uuid4().hex+".xlsx"
            wb.save(File_UPLOAD_PATH+os.sep+newfilename)
            os.system("del "+File_UPLOAD_PATH+os.sep+"*.png")
            return send_from_directory(File_UPLOAD_PATH,
                                       newfilename, as_attachment=True)
        else:
            return render_template("QRCode/qrcode.html",form=excel_form,msg="文件不符合要求")

    return render_template("QRCode/qrcode.html",form=excel_form)